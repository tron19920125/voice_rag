"""
文档处理模块
支持PDF/Word/Excel文档解析和QA数据解析
"""

import os
import re
import json
from typing import List, Dict, Optional
from pathlib import Path
import hashlib

# 文档解析库
try:
    from PyPDF2 import PdfReader
except ImportError:
    PdfReader = None

try:
    from docx import Document as DocxDocument
except ImportError:
    DocxDocument = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

import pandas as pd


class DocumentProcessor:
    """文档处理器：处理PDF/Word/Excel"""

    def __init__(self, chunk_size: int = 800, chunk_overlap: int = 100):
        """
        Args:
            chunk_size: 每个块的目标字符数
            chunk_overlap: 块之间的重叠字符数
        """
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap

    def process_file(self, file_path: str) -> List[Dict]:
        """
        处理文档文件

        Args:
            file_path: 文件路径

        Returns:
            文档块列表
        """
        ext = Path(file_path).suffix.lower()
        filename = Path(file_path).name

        if ext == ".pdf":
            text = self._extract_pdf(file_path)
        elif ext in [".docx", ".doc"]:
            text = self._extract_word(file_path)
        elif ext in [".xlsx", ".xls"]:
            text = self._extract_excel(file_path)
        else:
            raise ValueError(f"不支持的文件类型: {ext}")

        # 分块
        chunks = self._chunk_text(text)

        # 生成文档块
        results = []
        for i, chunk in enumerate(chunks):
            doc_id = self._generate_id(filename, i)
            results.append({
                "id": doc_id,
                "type": "document",
                "content": chunk,
                "title": filename,
                "question": None,
                "answer": None,
                "metadata": {
                    "source": filename,
                    "chunk_index": i,
                    "total_chunks": len(chunks)
                }
            })

        return results

    def _extract_pdf(self, file_path: str) -> str:
        """提取PDF文本"""
        if PdfReader is None:
            raise ImportError("请安装PyPDF2: uv add pypdf2")

        text_parts = []
        with open(file_path, "rb") as f:
            reader = PdfReader(f)
            for page in reader.pages:
                text_parts.append(page.extract_text())

        return "\n\n".join(text_parts)

    def _extract_word(self, file_path: str) -> str:
        """提取Word文本"""
        if DocxDocument is None:
            raise ImportError("请安装python-docx: uv add python-docx")

        doc = DocxDocument(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)

    def _extract_excel(self, file_path: str) -> str:
        """提取Excel文本（合并所有sheet）"""
        if load_workbook is None:
            raise ImportError("请安装openpyxl: uv add openpyxl")

        wb = load_workbook(file_path, read_only=True)
        text_parts = []

        for sheet in wb.worksheets:
            sheet_data = []
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    sheet_data.append(row_text)

            if sheet_data:
                text_parts.append(f"[Sheet: {sheet.title}]\n" + "\n".join(sheet_data))

        return "\n\n".join(text_parts)

    def _chunk_text(self, text: str) -> List[str]:
        """
        按语义分块（保留段落完整性）

        策略：
        1. 按段落分割
        2. 合并短段落达到目标chunk_size
        3. 拆分超长段落
        """
        # 按段落分割（双换行或单换行）
        paragraphs = re.split(r'\n\s*\n|\n', text)
        paragraphs = [p.strip() for p in paragraphs if p.strip()]

        chunks = []
        current_chunk = []
        current_length = 0

        for para in paragraphs:
            para_len = len(para)

            # 如果单个段落超过chunk_size，需要拆分
            if para_len > self.chunk_size * 1.5:
                # 先保存当前chunk
                if current_chunk:
                    chunks.append("\n\n".join(current_chunk))
                    current_chunk = []
                    current_length = 0

                # 按句子拆分超长段落
                sentences = re.split(r'([。！？.!?])', para)
                sentences = ["".join(sentences[i:i+2]) for i in range(0, len(sentences)-1, 2)]

                sub_chunk = []
                sub_length = 0
                for sent in sentences:
                    if sub_length + len(sent) > self.chunk_size and sub_chunk:
                        chunks.append("".join(sub_chunk))
                        # 保留重叠
                        sub_chunk = sub_chunk[-1:]
                        sub_length = len(sub_chunk[0]) if sub_chunk else 0

                    sub_chunk.append(sent)
                    sub_length += len(sent)

                if sub_chunk:
                    chunks.append("".join(sub_chunk))

            # 普通段落
            elif current_length + para_len > self.chunk_size and current_chunk:
                # 达到阈值，保存当前chunk
                chunks.append("\n\n".join(current_chunk))
                # 保留最后一个段落作为重叠
                if self.chunk_overlap > 0 and current_chunk:
                    current_chunk = [current_chunk[-1]]
                    current_length = len(current_chunk[0])
                else:
                    current_chunk = []
                    current_length = 0

                current_chunk.append(para)
                current_length += para_len
            else:
                # 继续累积
                current_chunk.append(para)
                current_length += para_len

        # 保存最后一个chunk
        if current_chunk:
            chunks.append("\n\n".join(current_chunk))

        return chunks

    def _generate_id(self, filename: str, index: int) -> str:
        """生成唯一ID"""
        content = f"{filename}_{index}"
        return f"doc_{hashlib.md5(content.encode()).hexdigest()[:12]}"


class QAProcessor:
    """QA数据处理器：处理JSON/JSONL/CSV/Excel/Markdown"""

    def process_file(self, file_path: str) -> List[Dict]:
        """
        处理QA文件

        Args:
            file_path: 文件路径

        Returns:
            QA记录列表
        """
        ext = Path(file_path).suffix.lower()
        filename = Path(file_path).name

        if ext == ".json":
            qa_list = self._parse_json(file_path)
        elif ext == ".jsonl":
            qa_list = self._parse_jsonl(file_path)
        elif ext == ".csv":
            qa_list = self._parse_csv(file_path)
        elif ext in [".xlsx", ".xls"]:
            qa_list = self._parse_excel(file_path)
        elif ext in [".md", ".markdown"]:
            qa_list = self._parse_markdown(file_path)
        else:
            raise ValueError(f"不支持的QA文件类型: {ext}")

        # 生成QA记录
        results = []
        for i, qa in enumerate(qa_list):
            qa_id = self._generate_id(filename, i)
            question = qa.get("question", "").strip()
            answer = qa.get("answer", "").strip()
            category = qa.get("category", "")

            if not question or not answer:
                continue  # 跳过空QA

            # Q+A拼接作为content
            content = f"{question}\n{answer}"

            results.append({
                "id": qa_id,
                "type": "qa",
                "content": content,
                "title": question[:50],  # 使用问题前50字作为标题
                "question": question,
                "answer": answer,
                "metadata": {
                    "source": filename,
                    "category": category
                }
            })

        return results

    def _parse_json(self, file_path: str) -> List[Dict]:
        """解析JSON文件（数组格式）"""
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        if isinstance(data, list):
            return data
        elif isinstance(data, dict) and "qa" in data:
            return data["qa"]
        else:
            return [data]

    def _parse_jsonl(self, file_path: str) -> List[Dict]:
        """解析JSONL文件（每行一个JSON对象）"""
        qa_list = []
        with open(file_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    qa_list.append(json.loads(line))
        return qa_list

    def _parse_csv(self, file_path: str) -> List[Dict]:
        """解析CSV文件"""
        df = pd.read_csv(file_path)
        return df.to_dict("records")

    def _parse_excel(self, file_path: str) -> List[Dict]:
        """解析Excel文件（第一个sheet）"""
        df = pd.read_excel(file_path)
        return df.to_dict("records")

    def _parse_markdown(self, file_path: str) -> List[Dict]:
        """
        解析Markdown FAQ格式

        格式：
        ## 问题1
        答案1

        ## 问题2
        答案2
        """
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()

        qa_list = []
        # 匹配 ## 标题 后的内容
        pattern = r'##\s+(.+?)\n(.*?)(?=\n##|\Z)'
        matches = re.findall(pattern, content, re.DOTALL)

        for question, answer in matches:
            question = question.strip()
            answer = answer.strip()
            if question and answer:
                qa_list.append({
                    "question": question,
                    "answer": answer,
                    "category": ""
                })

        return qa_list

    def _generate_id(self, filename: str, index: int) -> str:
        """生成唯一ID"""
        content = f"{filename}_{index}"
        return f"qa_{hashlib.md5(content.encode()).hexdigest()[:12]}"


def process_file(file_path: str) -> List[Dict]:
    """
    统一接口：自动识别文件类型并处理

    Args:
        file_path: 文件路径

    Returns:
        处理后的记录列表（文档块或QA对）
    """
    ext = Path(file_path).suffix.lower()

    # 文档类型
    if ext in [".pdf", ".docx", ".doc"]:
        processor = DocumentProcessor()
        return processor.process_file(file_path)

    # Excel特殊处理：检查是否为QA格式
    elif ext in [".xlsx", ".xls"]:
        # 尝试作为QA解析
        try:
            df = pd.read_excel(file_path, nrows=1)
            if "question" in df.columns and "answer" in df.columns:
                processor = QAProcessor()
                return processor.process_file(file_path)
        except Exception:
            pass

        # 否则作为文档
        processor = DocumentProcessor()
        return processor.process_file(file_path)

    # QA类型
    elif ext in [".json", ".jsonl", ".csv", ".md", ".markdown"]:
        processor = QAProcessor()
        return processor.process_file(file_path)

    else:
        raise ValueError(f"不支持的文件类型: {ext}")
